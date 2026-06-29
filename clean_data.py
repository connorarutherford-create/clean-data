#!/usr/bin/env python3
"""clean-data — Automated data cleaning. Handles date normalization,
phone formatting, currency stripping, deduplication, and more."""

import argparse
import csv
import os
import re
import sys
from datetime import datetime

# ── Security ───────────────────────────────────────────────────────────────────
MAX_FILE_SIZE = 50 * 1024 * 1024
ALLOWED_EXTS = {".csv", ".xlsx", ".xls"}

def check(path):
    path = os.path.abspath(os.path.expanduser(path))
    if not os.path.exists(path):
        sys.exit(f"ERROR: File not found: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext not in ALLOWED_EXTS:
        sys.exit(f"ERROR: Unsupported type '{ext}'. Use: {', '.join(ALLOWED_EXTS)}")
    if os.path.getsize(path) > MAX_FILE_SIZE:
        sys.exit(f"ERROR: File too large ({os.path.getsize(path)/1024/1024:.1f} MB). Max: 50 MB")
    if ext == ".csv":
        with open(path, "rb") as f:
            try: f.read(4096).decode("utf-8")
            except: sys.exit("ERROR: CSV contains non-text data.")
    return path

# ── Cleaners ──────────────────────────────────────────────────────────────────

def norm_date(val):
    if not val or not str(val).strip(): return ""
    v = str(val).strip()
    for f in ["%m/%d/%Y","%m/%d/%y","%Y-%m-%d","%d/%m/%Y","%m-%d-%Y","%Y/%m/%d",
              "%d-%m-%Y","%B %d, %Y","%b %d, %Y","%b %d %Y","%d %B %Y","%Y%m%d","%m.%d.%Y"]:
        try: return datetime.strptime(v, f).strftime("%Y-%m-%d")
        except: pass
    return v

def norm_num(val):
    if not val or not str(val).strip(): return ""
    v = str(val).strip().replace("$","").replace("€","").replace("£","").replace(",","")
    if v.endswith("%"):
        try: return str(float(v[:-1])/100)
        except: return v
    try: return str(float(v))
    except: return v

def norm_phone(val):
    if not val or not str(val).strip(): return ""
    d = re.sub(r"\D","",str(val))
    if len(d)==10: return f"({d[:3]}) {d[3:6]}-{d[6:]}"
    if len(d)==11 and d[0]=="1": return f"({d[1:4]}) {d[4:7]}-{d[7:]}"
    return val

def norm_zip(val):
    if not val or not str(val).strip(): return ""
    d = re.sub(r"\D","",str(val))
    if len(d)==9: return f"{d[:5]}-{d[5:]}"
    if len(d)==5: return d
    return val

def smart_clean(val, col=""):
    v = str(val).strip() if val else ""
    c = col.lower()
    if "date" in c: return norm_date(v)
    if any(x in c for x in ["phone","fax","mobile","cell"]): return norm_phone(v)
    if any(x in c for x in ["zip","postal","post code"]): return norm_zip(v)
    if any(x in c for x in ["price","cost","total","amount","revenue","salary","rate"]): return norm_num(v)
    return v

# ── File I/O ──────────────────────────────────────────────────────────────────

def load(path):
    path = check(path)
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            r = csv.DictReader(f); return list(r), r.fieldnames
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(h,[str(c) if c is not None else "" for c in r])) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows, h

def save(rows, fields, path):
    if path.endswith(".csv"):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(rows)
    else:
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(fields)
        for r in rows: ws.append([r.get(h,"") for h in fields])
        wb.save(path)
    print(f"✅ Saved: {path} ({len(rows)} rows)")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="Clean messy data files")
    p.add_argument("input"); p.add_argument("-o","--output")
    p.add_argument("--dedup", action="store_true"); p.add_argument("--auto", action="store_true")
    p.add_argument("--summary", action="store_true")
    args = p.parse_args()

    if not args.output:
        b, e = os.path.splitext(args.input); args.output = f"{b}_clean{e}"

    print(f"📂 Loading: {args.input}")
    rows, fields = load(args.input)
    print(f"   {len(rows)} rows, {len(fields)} columns: {', '.join(fields)}")

    changes = 0
    cleaned = []
    for row in rows:
        cr = {}
        for c in fields:
            v = row.get(c,"")
            if args.auto: cr[c] = smart_clean(v, c)
            else: cr[c] = v.strip() if v else ""
            if cr[c] != str(row.get(c,"")).strip(): changes += 1
        cleaned.append(cr)

    before = len(cleaned)
    if args.dedup or args.auto:
        seen = set(); deduped = []
        for r in cleaned:
            k = tuple(r.get(c,"") for c in fields)
            if k not in seen: seen.add(k); deduped.append(r)
        cleaned = deduped

    save(cleaned, fields, args.output)

    if args.summary or args.auto:
        print(f"\n📊 Summary: {len(rows)} in → {len(cleaned)} out, {changes} cells cleaned")

if __name__ == "__main__":
    main()
